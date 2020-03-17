import openpyxl
from datetime import datetime
from datetime import timedelta
from openpyxl.styles import Font, colors, Alignment
from prettytable import PrettyTable
import json


with open('setting.conf', 'r', encoding='utf-8') as f:
    conf = f.read()
    INFO_JSON = json.loads(conf)
try:
    FILENAME = INFO_JSON['filename']
    PATH = INFO_JSON['path']
    SYSNAME = INFO_JSON['sysname']
    PLANNAME = INFO_JSON['planname']
except Exception as e:
    input("应用配置信息有误！！")


def getDays(num):
    days = (datetime.now() + timedelta(days=num)).strftime("%Y/%m/%d")
    return days


class ReviewSys(object):
    def __init__(self):
        self.today = datetime.now().strftime("%Y/%m/%d")
        self.head = PrettyTable([SYSNAME])
        self.head.add_row(["1.新学知识"])
        self.head.add_row(["2.错题记录"])
        self.head.add_row(["3.今日复习"])
        self.head.add_row(["4.错题改正"])
        self.wb = openpyxl.load_workbook(PATH + FILENAME + '.xlsx')  # 打开excel文件
        self.all_sheets = self.wb.sheetnames  # 获得所有 sheets
        self.second_day = getDays(1)  # 第二天天后，第一次复习
        self.forth_day = getDays(3)  # 第四天，第二次复习
        self.seventh_day = getDays(6)  # 第七天，第三次复习
        self.fifth_day = getDays(14)  # 第十五天，第四次复习
        self.sexth_day = getDays(4)
        self.third_day = getDays(2)
        self.reviews = list()

    def change_font_math_knowlege(self, ws):
        sum_rows = ws.max_row
        sum_columns = ws.max_column
        for i in range(1, sum_columns + 1):
            ws.cell(row=sum_rows, column=i).alignment = Alignment(horizontal='center', vertical='center')
            if i == 3:
                font_style = Font(name='等线', color=colors.RED)
                ws.cell(row=sum_rows, column=i).font = font_style
            if i == 4:
                font_style = Font(name='等线', color=colors.DARKYELLOW)
                ws.cell(row=sum_rows, column=i).font = font_style
            if i == 5:
                font_style = Font(name='等线', color=colors.BLACK)
                ws.cell(row=sum_rows, column=i).font = font_style
            if i == 6:
                font_style = Font(name='等线', color=colors.BLUE)
                ws.cell(row=sum_rows, column=i).font = font_style

        self.wb.save(PATH + FILENAME + '.xlsx')

    def change_font_1800(self, ws):
        sum_rows = ws.max_row
        sum_columns = ws.max_column
        for i in range(1, sum_columns + 1):
            ws.cell(row=sum_rows, column=i).alignment = Alignment(horizontal='center', vertical='center')
            if i == 4 and ws.cell(row=sum_rows, column=4).value == "否":
                font_style = Font(name='等线', color=colors.DARKRED)
                ws.cell(row=sum_rows, column=i).font = font_style
            if i == 5:
                font_style = Font(name='等线', color=colors.RED)
                ws.cell(row=sum_rows, column=i).font = font_style
            if i == 6:
                font_style = Font(name='等线', color=colors.BLUE)
                ws.cell(row=sum_rows, column=i).font = font_style
        self.wb.save(PATH + FILENAME + '.xlsx')

    def save_kpint(self):
        ws = self.wb[self.all_sheets[0]]
        content = input('请输入学习内容：')
        ws.append([self.today, content, self.second_day, self.forth_day, self.seventh_day, self.fifth_day])
        self.change_font_math_knowlege(ws)

    def seve_mistakes(self):
        ws = self.wb[self.all_sheets[1]]
        q_num = input('题号：')
        knowlege = input('知识点：')
        y_n = '否'
        ws.append([q_num, knowlege, self.today, y_n, self.third_day, self.sexth_day])
        self.change_font_1800(ws)

    def print_reviws(self):
        table = PrettyTable(['编号', '复习内容'])
        for i, num in enumerate(self.reviews):
            table.add_row([str(i), str(num)])
        print(table)

    def see_plan(self):
        ws = self.wb[self.all_sheets[0]]
        for row in range(1, ws.max_row + 1):
            if ws.cell(row, 3).value == str(self.today) or ws.cell(row, 4).value == str(self.today) or ws.cell(row,
                                                                                                     5).value == str(
                    self.today) or ws.cell(row, 6).value == str(self.today):
                content = '知识点：' + ws.cell(row, 2).value
                if content not in self.reviews:
                    self.reviews.append(content)
        ws = self.wb[self.all_sheets[1]]
        for row in range(1, ws.max_row + 1):
            if (ws.cell(row, 5).value == str(self.today) or ws.cell(row, 6).value == str(self.today)) and ws.cell(row,
                                                                                                        4).value == "否":
                content = "习题：" + ws.cell(row, 1).value
                if ws.cell(row, 6).value == str(self.today):
                    content += "   [这是第二次做这道题了。]"
                if content not in self.reviews:
                    self.reviews.append(content)

    def change_mistakes(self, misktake):
        ws = self.wb[self.all_sheets[1]]
        for row in range(1, ws.max_row+1):
            if ws.cell(row, 1).value == misktake:
                ws.cell(row, 4).value = "会"
                ws.cell(row, 4).font = Font(name='等线', color="dd7545")
        self.wb.save(PATH + FILENAME + '.xlsx')


    def run(self):
        while(True):
            print(self.head)
            choice = input('请输入学习内容：')
            if choice == '1':
                self.save_kpint()
                self.save_plans()
            elif choice == '2':
                self.seve_mistakes()
                self.save_plans()
            elif choice == '3':
                self.see_plan()
                self.print_reviws()
                self.save_plans()
            elif choice == '4':
                """错题改正"""
                mistakes = input("请输入错题的题号：")
                self.change_mistakes(mistakes)
            elif choice == 'q' or choice == 'quit':
                self.save_plans()
                print("再见")
                break
            else:
                self.save_plans()

    def save_plans(self):
        with open(PLANNAME + '.txt', 'w+', encoding='utf-8') as fp:
            fp.write("复习计划：\n")
            fp.write("\n")
        self.see_plan()
        with open(PLANNAME+'.txt', 'a+', encoding='utf-8') as fp:
            fp.write(self.today + "的复习计划：\n")
            for i, content in enumerate(self.reviews):
                fp.write("\t" + str(i) + "\t" + str(content) + "\n")
            fp.write("\n")
            fp.write("\n")
        self.today = self.second_day
        self.reviews = []
        self.see_plan()
        with open(PLANNAME+'.txt', 'a+', encoding='utf-8') as fp:
            fp.write(self.today + "的复习计划：\n")
            for i, content in enumerate(self.reviews):
                fp.write("\t" + str(i) + "\t" + str(content) + "\n")
        self.today = datetime.now().strftime("%Y/%m/%d")
        self.reviews = []


if __name__ == '__main__':
    a = ReviewSys()
    a.run()
